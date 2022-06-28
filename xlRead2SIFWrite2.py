import networkx as nx
import plotly.graph_objects as go
from courseclass import Course
import matplotlib.pyplot as plt
import openpyxl


# sheet1 = "MathOnly4networkTesting"
# TODO: DONE: finish math dept prereq entry in SIF-readable CSV format (IN DELTESTEXP.xlsx)
# TODO: then to all depts(big)
# wb_obj = openpyxl.load_workbook(cprqfile)
# mathSheet = wb_obj[wb_obj.sheetnames[0]]

def xl2SIFnetworkcreator(xlWbFilePath,sheetIndex,startingRowOfEdgeEntries,columnNumofEdgeEntries,outputSIFnameAndOrPath):
    myWB = openpyxl.load_workbook(xlWbFilePath)
    mySheet = myWB[myWB.sheetnames[sheetIndex]]
    with open(outputSIFnameAndOrPath, 'w') as myOutFile:
        glbOrCtr = 0  # can handle up to 100 or nodes per course
        glbAndCtr = 0  # ^
        for i in range(startingRowOfEdgeEntries, mySheet.max_row):
            # while True:
            #     i = 49
            prContents = str(mySheet.cell(row=i, column=columnNumofEdgeEntries).value).strip().split(",")
            if str(prContents) != "['None']":
                orDict = {}  # RESETS FOR EACH COURSE, AVOIDING OVERLAPS
                andDict = {}  # ^
                for ea in prContents:
                    print("cont " + str(ea))
                    # handle OR node renaming
                    if "%" in ea:
                        curOR_ID = ea[ea.find("%"):ea.find("%") + 3]
                        print("cur " + str(curOR_ID))
                        if curOR_ID not in orDict:
                            # add new mapping {line% : global%} in orDict
                            if glbOrCtr >= 10:
                                orDict[curOR_ID] = "or" + str(glbOrCtr)
                            elif glbOrCtr < 10:
                                orDict[curOR_ID] = "or0" + str(glbOrCtr)
                            # increment orCounter
                            glbOrCtr += 1
                            print("ctr" + str(glbOrCtr))

                        if "%" in ea[3:]:  # if OR->OR, also remap second instance
                            print("2nd OR found")
                            subs = ea[3:]
                            curOR_ID2 = subs[subs.find("%"):subs.find("%") + 3]
                            print(curOR_ID2)
                            if curOR_ID2 not in orDict:
                                # add new mapping {line% : global%} in orDict
                                if glbOrCtr >= 10:
                                    orDict[curOR_ID2] = "or" + str(glbOrCtr)
                                elif glbOrCtr < 10:
                                    orDict[curOR_ID2] = "or0" + str(glbOrCtr)
                                # increment orCounter
                                glbOrCtr += 1
                                print("ctr" + str(glbOrCtr))

                        # EITHER WAY set/replace to global mapping (new or existing mapping)
                        ea = ea.replace(curOR_ID, orDict[curOR_ID])
                        ea = ea.replace(curOR_ID2, orDict[curOR_ID2])

                    if "&" in ea:
                        curAND_ID = ea[ea.find("&"):ea.find("&") + 3]
                        print("cur " + str(curAND_ID))
                        if curAND_ID not in andDict:
                            # add new mapping {line% : global%} in andDict
                            if glbAndCtr >= 10:
                                andDict[curAND_ID] = "and" + str(glbAndCtr)
                            elif glbAndCtr < 10:
                                andDict[curAND_ID] = "and0" + str(glbAndCtr)
                            # increment andCounter
                            glbAndCtr += 1
                            print("ctr" + str(glbAndCtr))

                        if "&" in ea[3:]:  # if AND->AND, also remap second instance
                            print("2nd AND found")
                            subs = ea[3:]
                            curAND_ID2 = subs[subs.find("&"):subs.find("&") + 3]
                            print(curAND_ID2)
                            if curOR_ID2 not in orDict:
                                # add new mapping {line% : global%} in andDict
                                if glbAndCtr >= 10:
                                    andDict[curAND_ID2] = "and" + str(glbAndCtr)
                                elif glbAndCtr < 10:
                                    andDict[curAND_ID2] = "and0" + str(glbAndCtr)
                                # increment andCounter
                                glbAndCtr += 1
                                print("ctr" + str(glbAndCtr))

                        # EITHER WAY set/replace to global mapping (new or existing mapping)
                        ea = ea.replace(curAND_ID, andDict[curAND_ID])
                        ea = ea.replace(curAND_ID2, andDict[curAND_ID2])

                    myOutFile.write(str(ea))
                    myOutFile.write('\n')
# TODO: change network display settings/etc
# TODO: auto-open in browser/etc, also make interactive (long term, using Dash-Cytoscape.js?)

cprqfile = "C:/Users/John DeForest/PycharmProjects/dartyclassdb1/deltestexp.xlsx"
xl2SIFnetworkcreator(cprqfile,0,2,12,'testExp.sif')
