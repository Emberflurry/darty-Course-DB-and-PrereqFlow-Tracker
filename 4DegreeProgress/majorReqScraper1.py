import openpyxl
import re


class Degree:
    def __init__(self, idnum, type, title, requirements=[]):
        self.idnum = int(idnum)
        self.type = str(type)
        self.title = str(title)
        self.requirements = requirements  # simple LoL to start? mayb pd df or pickle in future?
        # ea requirement 'entry' has a reqType, #ofReq, list2ChooseFrom, notAllowed/conditionals, countrestrictions
        # todo: should i make a requirements special class? Alec?
        # note: ^ I did. see below.

    def __str__(self):
        return str(self.idnum) + "_" + self.type + "_" + self.title


class Requirement:
    def __init__(self, rType, rQty, courses2choose, notAllowed, ctRestrict):
        self.rType = rType
        self.rQty = rQty
        self.courses2choose = courses2choose
        self.notAllowed = notAllowed
        self.ctRestrict = ctRestrict

    def __str__(self):
        return str(self.rType) + "_" + str(self.rQty) + "_" + str(self.courses2choose) + "_not:" + str(
            self.notAllowed) + "res:" + str(self.ctRestrict)


# file is excel file to load from
# sheet is name of the sheet in the file
def get_degrees_data(file, sheet):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]

    degrees = dict()  # id# : Degree object

    # Starts at 2 to skip title row and b/c openpyxl is 1-based instead of 0-based  --thank you Alec
    for i in range(2, sheet.max_row + 1):

        # Col 1 is degree type (maJor, miNor, Modified)
        degType = str(sheet.cell(row=i, column=1).value)
        # Col 2 is requirement type (prereq, need, rec)
        reqType = str(sheet.cell(row=i, column=2).value)
        # Col 3 is the degree ID (0 - ...)
        degID = int((sheet.cell(row=i, column=3).value))
        # Col 4 is the degree name/title
        degTitle = str((sheet.cell(row=i, column=4).value))
        # Col 5 is the number of courses required fulfilled for this requirement entry
        reqQty = int((sheet.cell(row=i, column=5).value))
        # Col 6 is the list of courses to pick from to satisfy the requirement entry
        reqCoursesEntry = ((sheet.cell(row=i, column=6).value))
        # Col 7 is the courses that are never allowed to fulfill that requirement entry
        reqCoursesNotAllowed = ((sheet.cell(row=i, column=7).value))
        # Col 8 contains count restrictions(ie: no more than 2 COSC courses can be used to fill this requirement entry)
        reqCountRestrictions = ((sheet.cell(row=i, column=3).value))

        newRequirement = Requirement(reqType, reqQty, reqCoursesEntry, reqCoursesNotAllowed, reqCountRestrictions)
        print(newRequirement)  # for shits and a few giggles

        if degID not in degrees:
            newDegree = Degree(degID, degType, degTitle)
            degrees[degID] = newDegree

        degrees[degID].requirements.append(newRequirement)

    return degrees


myFile = "C:/Users/John DeForest/Desktop/MATH major requirements.xlsx"
mySheetName = "compReadable"

outputDegreesDict = get_degrees_data(myFile, mySheetName)

print(outputDegreesDict)
for item in outputDegreesDict:
    print(outputDegreesDict[item])  # note: WORKS!
