import openpyxl
import dash
import dash_cytoscape as dcyto
from dash import html as dhtml
from dash import dcc  # ^ use this instead
from dash.dependencies import Input, Output, State
import re
import json
from queue import Queue

# import dash_core_components as dcc # PHASED OUT DO NOT USE THIS IMPORT SYNTAX
#suppress_callback_exceptions=True
dcyto.load_extra_layouts()

global resetNodeSelection
resetNodeSelection=False

def xl2SIFnetworkcreator(xlWbFilePath, sheetIndex, startingRowOfEdgeEntries, columnNumofEdgeEntries,
                         columnNumofNodeNames, columnNumofTitle, columnNumofDesc,
                         outputSIFnameAndOrPath):
    myWB = openpyxl.load_workbook(xlWbFilePath)
    mySheet = myWB[myWB.sheetnames[sheetIndex]]
    with open(outputSIFnameAndOrPath, 'w') as myOutFile:
        glbOrCtr = 0  # can handle note: INCREASED TO 0-999 (THREE DIGITS CAPACITY FOR ALL GLOBAL ORS/ANDS)
        glbAndCtr = 0  # ^           # note: Only use now if counting # of ors and ands for debugging...

        # FOR outputs to node/edge lists for Dash-Cyto graph visuals, see lines 102-119
        edgeList = []
        nodeList = []

        cumNodeSet = set()  # note: for keeping track of NODES to add/already been added
        lineDuplicateSet = set()  # note: makes sure repeated/duplicate EDGES are not added

        # note: adding nodes from the EDGELISTS
        for i in range(startingRowOfEdgeEntries, mySheet.max_row):
            prContents = str(mySheet.cell(row=i, column=columnNumofEdgeEntries).value).strip().split(",")

            if str(prContents) != "['None']":

                for ea in prContents:
                    if str(ea) not in lineDuplicateSet:
                        lineDuplicateSet.add(ea)

                        if "%" in ea:  # NOTE: the id of the OR nodes will be the %### but ill make the labels all OR s
                            glbOrCtr += 1
                            if "%" in ea[4:]:  # if OR->OR, also remap second instance 3: for 2dig, 4: for 3dig?
                                glbOrCtr += 1

                        if "&" in ea:
                            glbAndCtr += 1
                            if "&" in ea[4:]:  # if AND->AND, also remap second instance
                                glbAndCtr += 1

                        # add the lines to the output files for manual checking
                        ea = ea.strip()  # clean
                        myOutFile.write(str(ea))
                        myOutFile.write('\n')

                        # split the line data, then add to the list complex  ( ('src','targ','L'), (etc), (etc)...)

                        a = ea.split(" ")
                        mySrc = a[0]
                        if len(a) == 3:
                            myEdgeLabel = a[1]
                            myTarg = a[2]
                            edgeList.append((mySrc, myTarg, myEdgeLabel))

                            if myTarg not in cumNodeSet:
                                c = (myTarg, "", "")
                                nodeList.append(c)
                                cumNodeSet.add(myTarg)

                        if mySrc not in cumNodeSet:
                            b = (mySrc, "", "")
                            nodeList.append(b)
                            cumNodeSet.add(mySrc)

        # note: now add FLOATER NODES - ones that have no prereqs/connections
        for k in range(startingRowOfEdgeEntries, mySheet.max_row):
            nodeInfo = str(mySheet.cell(row=k, column=columnNumofNodeNames).value).strip()
            courseTitle = str(mySheet.cell(row=k, column=columnNumofTitle).value).strip()
            courseDescription = str(mySheet.cell(row=k, column=columnNumofDesc).value).strip()
            # TODO: ALSO ADD the OG PREREQ INFO (TEXT) SO THAT USERS CAN INTERPRET COMPLEXITIES
            # AND OFC the URL(SEE recent ORC scrape xlsx in Webscrapers folder and corresponding script-
            # ^integrate into final version),
            # coreqs,profs, etc -all that can be found from orc page...
            # maybe later also link relevant DEPARTMENT pages but that might be manual

            if "." not in nodeInfo:
                match = re.match(r"([a-z]+)([0-9]+)", nodeInfo, re.I)
                if match:
                    splitNodeInfo = match.groups()
                    # print(splitNodeInfo)

                    if len(splitNodeInfo[1]) == 1:
                        newNumStr = "00" + splitNodeInfo[1]
                    elif len(splitNodeInfo[1]) == 2:
                        newNumStr = "0" + splitNodeInfo[1]
                    else:
                        newNumStr = splitNodeInfo[1]

                    floaterNode = splitNodeInfo[0] + newNumStr
                else:
                    print("regex error in finding match of char/number boundary")

            elif "." in nodeInfo:
                floaterNode = nodeInfo

            newNodeTuple = floaterNode, courseTitle, courseDescription
            print("new", newNodeTuple)

            duplicateReplaced = False
            for g in range(0, len(nodeList) - 1):
                if floaterNode == nodeList[g][0]:
                    nodeList[g] = newNodeTuple
                    duplicateReplaced = True
                    break
            if duplicateReplaced is False:
                nodeList.append(newNodeTuple)

    # print(nodeList)
    # print(edgeList)
        for y in range(0, len(nodeList)-1):
            if nodeList[y][0][0] == "%":
                newOrTitle = "OR - 1 of child nodes is required to be fulfilled as prerequisite"
                nodeList[y] = nodeList[y][0], newOrTitle, "just an OR node"
            if nodeList[y][0][0] == "&":
                newAndTitle = "AND - all children are required to be fulfilled as prerequisite"
                nodeList[y] = nodeList[y][0], newAndTitle, "just an AND node"

    return nodeList, edgeList


cprqfile = "C:/Users/John DeForest/PycharmProjects/dartyclassdb1/2IntermediateProcessing/xlDBcleaning/deleteTestExportCURRENT3.xlsx"

myNodesLoL, myEdgesLoL = xl2SIFnetworkcreator(cprqfile, 0, 2, 7, 1, 2, 5,
                                              'edgelistOutput3.txt')  # 2nd param: 0 for MATH, 1 for MATH+ENGS
# note: this txt writing step^ is for manual checking of the reading from excel process,
#  really can just write straight to list format (as is DONE by the fn)

# print("--")
# print(myNodesLoL)
# print(myEdgesLoL)
# quit()

myApp = dash.Dash(__name__)

myNodes = [
    {'data': {'id': shortID, 'title': labelID1, 'desc': descText}, }
    for shortID, labelID1, descText in myNodesLoL
]
myEdges = [
    {'data': {'source': sourceID, 'target': targetID, 'label': labelID2}}
    for sourceID, targetID, labelID2 in myEdgesLoL
]

# print(myNodes)
# print(myEdges)

myAllElements = myNodes + myEdges

myDefaultStylesheet = [
    {'selector': 'node', 'style': {'label': 'data(id)'}},
    # {'selector': 'edge', 'style': {'label': 'data(label)'}},
    {'selector': 'edge', 'style': {'curve-style': 'bezier'}},
    {'selector': 'edge', 'style': {'mid-target-arrow-color': 'blue',
                                   'mid-target-arrow-shape': 'vee',
                                   'line-color': 'grey', 'arrow-scale': 3.5, }},
    {"selector": 'node[id = "CLER001"]','style': {'shape': 'rectangle', 'width': 75,'height':40}}
]

sidebarStyles = {
    'tab1': {'height': 'calc(98vh-115px)'},
    'contentStyle1': {'overflow-y': 'scroll', 'height': 'calc(50%-25px)','border':'thin lightgrey solid'}
}

#note: global stuff:
myCyto_id = 'cytoscape-event-callbacks-2'

print("defined app")
myApp.layout = dhtml.Div([

    #note: this div for cyto node layout
    dhtml.Div(className='eight columns', children=[
        # dropdown layou menu
        # dcc.Dropdown(
        #     id='dropdown-update-layout', value='dagre', clearable=False,
        #     options=[
        #         {'label': myLayoutSelName, 'value': myLayoutSelName} for myLayoutSelName in
        #         ['dagre', 'breadthfirst', 'klay', 'euler']
        #     ]
        # ),
        dcyto.Cytoscape(
            # id='cytoscape'
            # id='cytoscape-update-layout',  # changed to this for dropdown menu ONLY
            id=myCyto_id,
            elements=myAllElements,
            stylesheet=myDefaultStylesheet,
            style={'width': '90%', 'height': '450px'},
            # layout choices:
            # ["random",  bad
            # "preset",   If you have preset node locations, good.
            # "circle",   not for this project
            # "concentric", not for this project
            # "grid",       not for this project
            # "breadthfirst",  pretty good
            # "cose",          too bunched
            # added extras:------------------
            # `cose-bilkent` can't use for some reason     https://github.com/cytoscape/cytoscape.js-cose-bilkent
            # `cola`         nasty                         https://github.com/cytoscape/cytoscape.js-cola
            # `euler`        crashes                       https://github.com/cytoscape/cytoscape.js-dagre
            # `spread`       BAD                           https://github.com/cytoscape/cytoscape.js-spread
            # `dagre`        pretty good                   https://github.com/cytoscape/cytoscape.js-dagre
            # `klay`         decent                        https://github.com/cytoscape/cytoscape.js-klay
            layout={'name': 'dagre',
                    'roots': '[id = "MATH001"]'}
        ),
        # dhtml.P(id='cytoscape-mouseoverNodeData-output'),
        dhtml.Blockquote(id='cytoscape-mouseoverNodeData-output')

        # dhtml.Caption(id='cytoscape-mouseoverNodeData-output') NO, BAD.
        # dhtml.Title(id='cytoscape-mouseoverNodeData-output') NO DOESNT WORK
    ]),

    #note: this div 4 sidebar info
    dhtml.Div(className='four columns', children=[
                # dcc.Textarea #TODO try this instead/other html content types next after testing 'Tabs' as below
                dcc.Tabs(id='tabs', children=[
                    dcc.Tab(label='Click on a node for full ORC content',  # label here is the TITLE of the info tab
                            children=[
                        dhtml.Div(style=sidebarStyles['tab1'], children=[
                            dhtml.P('Node Data:'),  # SUBTITLE just above content
                            dhtml.Pre(
                                id='tap-node-data-output1',
                                style=sidebarStyles['contentStyle1']
                            )
                        ])
                    ]),

                    #note: adding color picker for edges i think
                    dcc.Tab(label='Control Panel1', children=[
                        #note: CANT FIND dash reusable components so using dcc.Input instead of drc.NamedInput
                        dcc.Input(name='inColor1',id='inputEdgeColor1',type='text',value='#0074D9'),
                        dcc.Input(name='outColor1',id='outputEdgeColor1',type='text',value='#FF4136')
                    ])

                ])
            ]),
    dhtml.Button('Clear Selection', id='clear-sel-button',n_clicks_timestamp=0),
    dhtml.Div(id='placeholder')  # TODO: WTF IS placeholder, no idea what this line is
])

def nodeBacktracerFull(rootNodeID):
    finalNodeSet=[]
    finalEdgeSet=[]
    frontierQ=Queue(maxsize=0)
    frontierQ.put(rootNodeID)
    while not frontierQ.empty():
        curNode=frontierQ.get()
        finalNodeSet.append(curNode)
        curChildren=[]
        for eaEdge in myEdges:
            if eaEdge['data']['target'] == curNode:
                finalEdgeSet.append(eaEdge)
                for eaNode in myNodes:
                    if eaNode['data']['id'] == eaEdge['data']['source'] and eaNode['data']['id'] not in finalNodeSet:
                        curChildren.append(eaNode['data']['id'])
        for c in curChildren:
            frontierQ.put(c)
    return finalNodeSet,finalEdgeSet  #note: NodeSet is list of str, EdgeSet is JSON OBJECTS!!

# for node click, draw per/post reqs
@myApp.callback(Output(myCyto_id, 'stylesheet'),
                [Input(myCyto_id, 'tapNodeData'),
                 Input('inputEdgeColor1', 'value'),
                 Input('outputEdgeColor1', 'value')])
def generate_stylesheet(node, incolor, outcolor):
    global resetNodeSelection
    print(resetNodeSelection)
    if not node:
        return myDefaultStylesheet
    if resetNodeSelection == True:
        resetNodeSelection = False
        return myDefaultStylesheet
    else:
        stylesheet = [{"selector": 'node',
            'style': {'opacity': 0.3,
            }
        },
            {'selector': 'edge',
            'style': {'opacity': 0.2, "curve-style": "bezier",
            }
        }, {"selector": 'node[id = "CLER001"]','style': {'shape': 'rectangle', 'width': 75,'height':40}},

            {#"selector": 'node[id = "{}"]'.format(node['data']['id']),
            "selector": '[title *= a]',  #TODO :DOES THIS WORK or see comment above or ePAIR SUAGE BELOW
            "style": {  #note: NEED TO KEEP THIS CHUNK EVEN IF IT DOESNT DO ANYTHING BC WEIRD ERROR
                # 'background-color': '#B10DC9',
                # "border-color": "purple",
                # "border-width": 2,
                # "border-opacity": 1,
                # "opacity": 1,
                #
                # "label": "data(label)",
                # "color": "#B10DC9",
                # "text-opacity": 1,
                # "font-size": 12,
                # 'z-index': 9999
            }
        }]

        #TODO: USE NEW ITERATIVE BACKTRACER, THEN ADD STYLESHEET STUFF FOR EDGES AND NODES PREREQS

        # (instead of the following: BUT KEEP THE FOLLOWING, maybe for future CHOICES OF PREREQS DISPLAY)
        for eaNode in myNodes:
            #print(eaNode)
            #print(node)
            if eaNode['data']['id'] == node['id']:
                stylesheet.append({
                    "selector": 'node[id = "{}"]'.format(eaNode['data']['id']),
                    "style": {
                        'background-color': '#B10DC9',
                        "border-color": "purple",
                        "border-width": 2,
                        "border-opacity": 1,
                        "opacity": 1,

                        "label": "data(label)",
                        "color": "#B10DC9",
                        "text-opacity": 1,
                        "font-size": 12,
                        'z-index': 9999
                    }
                })
        for ePair in myEdges:
            # print("edge / source / target")
            # print(ePair)
            # print(ePair['data']['source'])
            # print(ePair['data']['target'])
            if ePair['data']['target'] == node['id']:
                stylesheet.append({
                    "selector": 'node[id = "{}"]'.format(ePair['data']['source']),  #TODO: MIGHT NOT WORK IDK
                    "style": {
                        'background-color': outcolor,
                        'opacity': 0.9
                    }
                })
                # stylesheet.append({
                #     "selector": 'edge[id= "{}"]'.format(ePair['id']),  #TODO: same concern as above
                #     "style": {
                #         "mid-target-arrow-color": outcolor,
                #         "mid-target-arrow-shape": "vee",
                #         "line-color": outcolor,
                #         'opacity': 0.9,
                #         'z-index': 5000
                #     }
                # })
            if ePair['data']['target'] == node['id']:  #TODO
                stylesheet.append({
                    "selector": 'node[id = "{}"]'.format(ePair['data']['source']),
                    "style": {
                        'background-color': incolor,
                        'opacity': 0.9,
                        'z-index': 9999
                    }
                })
                # stylesheet.append({
                #     "selector": 'edge[id= "{}"]'.format(ePair['id']),  #TODO
                #     "style": {
                #         "mid-target-arrow-color": incolor,
                #         "mid-target-arrow-shape": "vee",
                #         "line-color": incolor,
                #         'opacity': 1,
                #         'z-index': 5000
                #     }
                # })
                #note: below lines are raw-copied from the demo example, above are my modified.

        # for edge in node['edgesData']:
        #     if edge['source'] == node['data']['id']:
        #         stylesheet.append({
        #             "selector": 'node[id = "{}"]'.format(edge['target']),
        #             "style": {
        #                 'background-color': outcolor,
        #                 'opacity': 0.9
        #             }
        #         })
        #         stylesheet.append({
        #             "selector": 'edge[id= "{}"]'.format(edge['id']),
        #             "style": {
        #                 "mid-target-arrow-color": outcolor,
        #                 "mid-target-arrow-shape": "vee",
        #                 "line-color": outcolor,
        #                 'opacity': 0.9,
        #                 'z-index': 5000
        #             }
        #         })
        #
        #     if edge['target'] == node['data']['id']:
        #         stylesheet.append({
        #             "selector": 'node[id = "{}"]'.format(edge['source']),
        #             "style": {
        #                 'background-color': incolor,
        #                 'opacity': 0.9,
        #                 'z-index': 9999
        #             }
        #         })
        #         stylesheet.append({
        #             "selector": 'edge[id= "{}"]'.format(edge['id']),
        #             "style": {
        #                 "mid-target-arrow-color": incolor,
        #                 "mid-target-arrow-shape": "vee",
        #                 "line-color": incolor,
        #                 'opacity': 1,
        #                 'z-index': 5000
        #             }
        #         })

        return stylesheet


# for node click, pulls node data
@myApp.callback(Output('tap-node-data-output1','children'),
                [Input(myCyto_id, 'tapNodeData')])
def displayTapNodeData(data):
    global resetNodeSelection
    # if data:
    #     if data['id'] == 'CLER001':
    #         resetNodeSelection = True
    return str(json.dumps(data, indent=2)).strip("{").strip("}")  # indent=2 ?


# for hover over node, pulls node name/title basic info
@myApp.callback(Output('cytoscape-mouseoverNodeData-output', 'children'),
                Input(myCyto_id, 'mouseoverNodeData'))
def displayTapNodeData(data):
    global resetNodeSelection
    if data:
        if data['id'] == 'CLER001':
            resetNodeSelection = True
        return "Hovered Node: " + data['id']+": "+data['title']
        # og was 'label', matches creation of the html/json Node List of Lists called myNodes (7/11/22)


# does dropdown layout menu stuff
# @myApp.callback(Output('cytoscape-update-layout', 'layout'),
#                 Input('dropdown-update-layout', 'value'))
# def update_layout(layout):
#     return {'name': layout, 'animate': True}


if __name__ == '__main__':
    myApp.run_server(debug=True)
