import openpyxl
import dash
import dash_cytoscape as dcyto
from dash import html as dhtml
from dash import dcc  # ^ use this instead
from dash.dependencies import Input, Output, State
import re
import json
from queue import Queue
import dash_bootstrap_components as dbc

# import dash_core_components as dcc # PHASED OUT DO NOT USE THIS IMPORT SYNTAX
# suppress_callback_exceptions=True
dcyto.load_extra_layouts()

global resetNodeSelection
resetNodeSelection = False

# global requisiteDisplaySetting  # pre, post, both
# requisiteDisplaySetting = "pre"

def xl2SIFnetworkcreator(xlWbFilePath, sheetIndex, startingRowOfEdgeEntries, columnNumofEdgeEntries,
                         columnNumofNodeNames, columnNumofTitle, columnNumofDesc,
                         outputSIFnameAndOrPath, PosSheetIndex):
    myWB = openpyxl.load_workbook(xlWbFilePath)
    mySheet = myWB[myWB.sheetnames[sheetIndex]]
    positionSheet = myWB[myWB.sheetnames[PosSheetIndex]]
    with open(outputSIFnameAndOrPath, 'w') as myOutFile:
        glbOrCtr = 0  # can handle note: INCREASED TO 0-999 (THREE DIGITS CAPACITY FOR ALL GLOBAL ORS/ANDS)
        glbAndCtr = 0  # ^           # note: Only use now if counting # of ors and ands for debugging...

        # FOR outputs to node/edge lists for Dash-Cyto graph visuals, see lines 102-119
        edgeList = []
        nodeList = list()

        cumNodeSet = set()  # note: for keeping track of NODES to add/already been added
        lineDuplicateSet = set()  # note: makes sure repeated/duplicate EDGES are not added

        # note: adding nodes from the EDGELISTS
        for i in range(startingRowOfEdgeEntries, mySheet.max_row):
            prContents = str(mySheet.cell(row=i, column=columnNumofEdgeEntries).value).strip().split(",")

            if str(prContents) != "['None']":

                for ea in prContents:
                    if str(ea) not in lineDuplicateSet:
                        lineDuplicateSet.add(ea)

                        if "%" in ea:
                            # NOTE: the id of the OR nodes will be the %### but
                            # note: ^i can make the labels all ORs later in styling
                            glbOrCtr += 1
                            if "%" in ea[4:]:  # if OR->OR, also remap second instance 3: for 2dig, 4: for 3dig?
                                glbOrCtr += 1

                        if "&" in ea:
                            glbAndCtr += 1
                            if "&" in ea[4:]:  # if AND->AND, also remap second instance
                                glbAndCtr += 1

                        # add the lines to the output files for manual checking
                        ea = ea.strip()  # clean
                        myOutFile.write(str(ea))
                        myOutFile.write('\n')

                        # split the line data, then add to the list complex  ( ('src','targ','L'), (etc), (etc)...)

                        a = ea.split(" ")
                        mySrc = a[0]
                        if len(a) == 3:
                            myEdgeLabel = a[1]
                            myTarg = a[2]
                            edgeList.append((mySrc, myTarg, myEdgeLabel))

                            if myTarg not in cumNodeSet:
                                c=list()
                                c.append(myTarg)
                                c.append("")
                                c.append("")
                                c.append("")
                                c.append("")
                                #c = [myTarg, "", ""]
                                nodeList.append(c)
                                cumNodeSet.add(myTarg)

                        if mySrc not in cumNodeSet:
                            b=list()
                            b.append(mySrc)
                            b.append("")
                            b.append("")
                            b.append("")
                            b.append("")
                            #b = [mySrc, "", ""]
                            nodeList.append(b)
                            cumNodeSet.add(mySrc)

        # note: now add FLOATER NODES - ones that have no prereqs or connections
        for k in range(startingRowOfEdgeEntries, mySheet.max_row):
            nodeInfo = str(mySheet.cell(row=k, column=columnNumofNodeNames).value).strip()
            courseTitle = str(mySheet.cell(row=k, column=columnNumofTitle).value).strip()
            courseDescription = str(mySheet.cell(row=k, column=columnNumofDesc).value).strip()
            # TODO:  M A N Y :
            # TODO: ALSO ADD the OG PREREQ INFO (RAW TEXT) SO THAT USERS CAN INTERPRET COMPLEXITIES
            # AND OFC the URL(SEE recent ORC scrape xlsx in Webscrapers folder and corresponding script-
            # ^integrate into final version),
            # coreqs,profs, etc -all that can be found from orc page...
            # maybe later also link relevant DEPARTMENT pages but that might be manual
            # can try to autolink LAYUPLIST pages for ea !!

            if "." not in nodeInfo:
                match = re.match(r'([a-z]+)([0-9]+)', nodeInfo, re.I)
                if match:
                    splitNodeInfo = match.groups()
                    # print(splitNodeInfo)

                    if len(splitNodeInfo[1]) == 1:
                        newNumStr = "00" + splitNodeInfo[1]
                    elif len(splitNodeInfo[1]) == 2:
                        newNumStr = "0" + splitNodeInfo[1]
                    else:
                        newNumStr = splitNodeInfo[1]

                    floaterNode = splitNodeInfo[0] + newNumStr
                elif not match:
                    print("regex error1 in finding match of char/number boundary w "+str(nodeInfo))

            elif "." in nodeInfo:
                floaterNode = nodeInfo

            initXpos, initYpos = 50, 50
            #newNodeTupleList = [floaterNode, courseTitle, courseDescription, initXpos, initYpos]
            newNodeTupleList = list()
            newNodeTupleList.append(floaterNode)
            newNodeTupleList.append(courseTitle)
            newNodeTupleList.append(courseDescription)
            newNodeTupleList.append(initXpos)
            newNodeTupleList.append(initYpos)

            # print("new", newNodeTupleList)

            duplicateReplaced = False
            for g in range(0, len(nodeList) - 1):
                if floaterNode == nodeList[g][0]:
                    nodeList[g] = newNodeTupleList
                    duplicateReplaced = True
                    break
            if duplicateReplaced is False:
                nodeList.append(newNodeTupleList)

        # print(nodeList)
        # print(edgeList)
        for y in range(0, len(nodeList) - 1):
            if nodeList[y][0][0] == "%":
                newOrTitle = "OR - 1 of child nodes is required to be fulfilled as prerequisite"
                nodeList[y] = [str(nodeList[y][0]), str(newOrTitle + "just an OR node"), "", "", ""]
            if nodeList[y][0][0] == "&":
                newAndTitle = "AND - all children are required to be fulfilled as prerequisite"
                nodeList[y] = [str(nodeList[y][0]), str(newAndTitle + "just an AND node"), "", "", ""]


        for i in range(2, positionSheet.max_row+1):
            nodeID = str(positionSheet.cell(row=i, column=1).value)
            #print(nodeID)
            Xpos = str(positionSheet.cell(row=i, column=2).value).strip()
            Ypos = str(positionSheet.cell(row=i, column=3).value).strip()

            #assuming no duplicates:
            bool = False
            for q in range(0, len(nodeList)-1):
                if str(nodeID) == str(nodeList[q][0]):
                    print(str(nodeID)+" // "+str(nodeList[q][0]))
                    #print(nodeList[q])
                    #print(nodeList[q][3])
                    nodeList[q][3] = Xpos
                    nodeList[q][4] = Ypos
                    bool = True

            if bool is False:
                print("node"+str(nodeID)+"didntLineUpWAnyCurrentNodes-no match for locationData pairing")
    # print(nodeList[3])
    # print(nodeList[53])
    # print(nodeList[13])
    # print(nodeList[23])
    # print(nodeList[78])
    # for itemm in nodeList:
    #     if str(itemm[4]) == '':
    #         print(str(itemm[0])+"missingCoords")
    return nodeList, edgeList


cprqfile = "C:/Users/John DeForest/PycharmProjects/dartyclassdb1/2IntermediateProcessing/xlDBcleaning/deleteTestExportCURRENT3.xlsx"

myNodesLoL, myEdgesLoL = xl2SIFnetworkcreator(cprqfile, 0, 2, 7, 1, 2, 5,
                                              'edgelistOutput3.txt', 1)  # 2nd param: 0 for MATH, 1 for MATH+ENGS
# note: this txt writing step^ is for manual checking of the reading from excel process,
#  really can just write straight to list format (AS IS DONE by the fn)

# print("--")
# print(myNodesLoL)  # this is a List of Lists, there is nothing funny about this. :|
# print(myEdgesLoL)
# quit()

myApp = dash.Dash(__name__, title='Darty Course-Flow Viz',external_stylesheets=[dbc.themes.BOOTSTRAP])

myNodes = [
    {'data': {'id': shortID, 'title': labelID1, 'desc': descText},
     'position': {'x': int(Xpos1), 'y': int(Ypos1)}}
    for shortID, labelID1, descText, Xpos1, Ypos1 in myNodesLoL
]
myEdges = [
    {'data': {'id': sourceID + targetID, 'source': sourceID, 'target': targetID, 'label': labelID2}}
    for sourceID, targetID, labelID2 in myEdgesLoL
]

# print(myNodes)
# print(myEdges)

myAllElements = myNodes + myEdges

myDefaultStylesheet = [
    {'selector': 'node',
     'style': {'label': 'data(id)', 'shape': 'rectangle', 'width': 100, 'height': 50, 'text-justification': 'center',
               'text-halign': 'center',
               'text-valign': 'center'}},
    # {'selector': 'edge', 'style': {'label': 'data(label)'}},
    {'selector': 'edge', 'style': {'curve-style': 'bezier'}},
    {'selector': 'edge', 'style': {'mid-target-arrow-color': 'blue',
                                   'mid-target-arrow-shape': 'vee',
                                   'line-color': 'grey', 'arrow-scale': 3.5, }},
    {"selector": 'node[id = "CLER001"]', 'style': {'shape': 'rectangle', 'width': 100, 'height': 60}}
]
orLabel = "OR"
myDefaultStylesheet.append({'selector': 'node[id *= "%"]',
                             'style': {'label': orLabel}})

sidebarStyles = {
    'tab1': {'height': 'calc(98vh-115px)'},
    'contentStyle1': {'overflow-y': 'scroll', 'height': 'calc(50%-25px)', 'border': 'thin lightgrey solid'}
}

# note: global stuff:
myCyto_id = 'cytoscape-event-callbacks-2'

print("defined app")
headerSection = [
    dhtml.H1(children='Dartmouth Course-Flow Vizualizer', style={'color': '#096A3F'}),
    dhtml.Div(children="course data from ORC as of 7/22, developed and maintained by John DeForest and visualized through a Plotly-Dash-Cytoscape config")
    ]
graphSection = [
    dhtml.Div(className='eight columns', children=[
            # dropdown layout menu
            # dcc.Dropdown(
            #     id='dropdown-update-layout', value='dagre', clearable=False,
            #     options=[
            #         {'label': myLayoutSelName, 'value': myLayoutSelName} for myLayoutSelName in
            #         ['dagre', 'breadthfirst', 'klay', 'euler']
            #     ]
            # ),
            dcyto.Cytoscape(
                # id='cytoscape'
                # id='cytoscape-update-layout',  # changed to this for dropdown menu ONLY
                id=myCyto_id,
                elements=myAllElements,
                stylesheet=myDefaultStylesheet,
                style={'width': '100%', 'height': '750px'},
                # layout choices:
                # ["random",  bad
                # "preset",   If you have preset node locations, good.
                # "circle",   not for this project
                # "concentric", not for this project
                # "grid",       not for this project
                # "breadthfirst",  pretty good
                # "cose",          too bunched
                # added extras:------------------
                # `cose-bilkent` can't use for some reason     https://github.com/cytoscape/cytoscape.js-cose-bilkent
                # `cola`         nasty                         https://github.com/cytoscape/cytoscape.js-cola
                # `euler`        crashes                       https://github.com/cytoscape/cytoscape.js-dagre
                # `spread`       BAD                           https://github.com/cytoscape/cytoscape.js-spread
                # `dagre`        pretty good                   https://github.com/cytoscape/cytoscape.js-dagre
                # `klay`         decent                        https://github.com/cytoscape/cytoscape.js-klay
                layout={'name': 'preset',
                        'roots': '[id = "MATH001"]'}
            ),
            # dhtml.P(id='cytoscape-mouseoverNodeData-output'),
            dhtml.Blockquote(id='cytoscape-mouseoverNodeData-output')

            # dhtml.Caption(id='cytoscape-mouseoverNodeData-output') NO, BAD.
            # dhtml.Title(id='cytoscape-mouseoverNodeData-output') NO DOESNT WORK
        ])
]
sidebarSection = [
    dhtml.Div(className='four columns', children=[
        # dcc.Textarea #TODO try this instead/other html content types next after testing 'Tabs' as below
        dcc.Tabs(id='tabs', children=[
            dcc.Tab(label='Click on a node for full ORC content',  # label here is the TITLE of the info tab
                    children=[
                        dhtml.Div(style=sidebarStyles['tab1'], children=[
                            dhtml.P('Node Data:'),  # SUBTITLE just above content
                            dhtml.Pre(
                                id='tap-node-data-output1',
                                style=sidebarStyles['contentStyle1']
                            )
                        ])
                    ]),

            # note: adding color picker for edges i think
            dcc.Tab(label='Control Panel', children=[
                # note: CANT FIND dash reusable components drc so using dcc.Input instead of drc.NamedInput, works
                dcc.Dropdown(['prerequisites', 'postrequisites', 'both', 'neither (why would you do this?)'], 'prerequisites', id='reqDropdown1'),
                #dcc.Input(name='inColor1', id='inputEdgeColor1', type='text', value='#0074D9'),
                #dcc.Input(name='outColor1', id='outputEdgeColor1', type='text', value='#FF4136'),
            ])

        ])
    ])]

nodeDataCard = dbc.Card(
    dbc.CardBody([
            dcc.Tabs(id='tabs', children=[
                        dcc.Tab(label='Click on a node for full ORC content',  # label here is the TITLE of the info tab
                                children=[
                                    dhtml.Div(style=sidebarStyles['tab1'], children=[
                                        dhtml.P('Node Data:'),  # SUBTITLE just above content
                                        dhtml.Pre(
                                            id='tap-node-data-output1',
                                            style=sidebarStyles['contentStyle1']
                                        )
                                    ])
                                ])]),
            dhtml.P("This is tab 1!", className="card-text"),
            dbc.Button("Click here", color="success"),
        ]),
    className="mt-3",)
ctrlPanelCard = dbc.Card(dbc.CardBody(
        [
            dhtml.P("This is tab 2!", className="card-text"),
            dbc.Button("Don't click here", color="danger"),

        ]
    ),
    className="mt-3",
)
# myApp.layout = dhtml.Div(
#     [
#         #headerSection,
#         # dbc.Row([
#         #     dbc.Col(dhtml.Div(children=graphSection), width=8),  # True "auto" 1-12  for expand, snugFit, gridWidthInteger resp.
#         #     dbc.Col(dhtml.Div(children=sidebarSection), width=4)
#         # ]),
#         dbc.Tabs([
#             dbc.Tab(nodeDataCard, label="Node Data"),
#             dbc.Tab(ctrlPanelCard, label="Display Controls")
#         ])
#     ]
# )
# myApp.layout = dbc.Tabs([
#             dbc.Tab(nodeDataCard, label="Node Data"),
#             dbc.Tab(ctrlPanelCard, label="Display Controls"),
#             dbc.Container(graphSection),
#             dbc.Container(sidebarSection),
#         ])
myApp.layout = dhtml.Div([
    dbc.Row([
        dbc.Col(dbc.Container(dbc.Col(children=graphSection)), width=9),
        dbc.Col(dbc.Container([
        dbc.Row(id="testID1", children=[dbc.Col(nodeDataCard)],),
        dbc.Row(dbc.Col(ctrlPanelCard)),
        dcc.Tab(label='Control Panel', children=[
                # note: CANT FIND dash reusable components drc so using dcc.Input instead of drc.NamedInput, works
                dcc.Dropdown(['prerequisites', 'postrequisites', 'both', 'neither (why would you do this?)'], 'prerequisites', id='reqDropdown1'),
                #dcc.Input(name='inColor1', id='inputEdgeColor1', type='text', value='#0074D9'),
                #dcc.Input(name='outColor1', id='outputEdgeColor1', type='text', value='#FF4136'),
            ])
    ], fluid=True), width=3)
    ]),

])



# myApp.layout = dhtml.Div([
#     dhtml.H1(children='Dartmouth Course-Flow Vizualizer', style={'color': '#096A3F'}),
#     dhtml.Div(children="course data from ORC as of 7/22, developed and maintained by John DeForest and visualized through a Plotly-Dash-Cytoscape config"),
#     # note: this div for cyto node layout
#     dhtml.Div(className='eight columns', children=[
#         # dropdown layout menu
#         # dcc.Dropdown(
#         #     id='dropdown-update-layout', value='dagre', clearable=False,
#         #     options=[
#         #         {'label': myLayoutSelName, 'value': myLayoutSelName} for myLayoutSelName in
#         #         ['dagre', 'breadthfirst', 'klay', 'euler']
#         #     ]
#         # ),
#         dcyto.Cytoscape(
#             # id='cytoscape'
#             # id='cytoscape-update-layout',  # changed to this for dropdown menu ONLY
#             id=myCyto_id,
#             elements=myAllElements,
#             stylesheet=myDefaultStylesheet,
#             style={'width': '90%', 'height': '450px'},
#             # layout choices:
#             # ["random",  bad
#             # "preset",   If you have preset node locations, good.
#             # "circle",   not for this project
#             # "concentric", not for this project
#             # "grid",       not for this project
#             # "breadthfirst",  pretty good
#             # "cose",          too bunched
#             # added extras:------------------
#             # `cose-bilkent` can't use for some reason     https://github.com/cytoscape/cytoscape.js-cose-bilkent
#             # `cola`         nasty                         https://github.com/cytoscape/cytoscape.js-cola
#             # `euler`        crashes                       https://github.com/cytoscape/cytoscape.js-dagre
#             # `spread`       BAD                           https://github.com/cytoscape/cytoscape.js-spread
#             # `dagre`        pretty good                   https://github.com/cytoscape/cytoscape.js-dagre
#             # `klay`         decent                        https://github.com/cytoscape/cytoscape.js-klay
#             layout={'name': 'dagre',
#                     'roots': '[id = "MATH001"]'}
#         ),
#         # dhtml.P(id='cytoscape-mouseoverNodeData-output'),
#         dhtml.Blockquote(id='cytoscape-mouseoverNodeData-output')
#
#         # dhtml.Caption(id='cytoscape-mouseoverNodeData-output') NO, BAD.
#         # dhtml.Title(id='cytoscape-mouseoverNodeData-output') NO DOESNT WORK
#     ]),
#
#     # note: this div 4 sidebar info
#     # TODO: MAKE INTO AN ACTUAL SIDEBAR, NOT THE BOTTOM TAB DISPLAY
#     dhtml.Div(className='four columns', children=[
#         # dcc.Textarea #TODO try this instead/other html content types next after testing 'Tabs' as below
#         dcc.Tabs(id='tabs', children=[
#             dcc.Tab(label='Click on a node for full ORC content',  # label here is the TITLE of the info tab
#                     children=[
#                         dhtml.Div(style=sidebarStyles['tab1'], children=[
#                             dhtml.P('Node Data:'),  # SUBTITLE just above content
#                             dhtml.Pre(
#                                 id='tap-node-data-output1',
#                                 style=sidebarStyles['contentStyle1']
#                             )
#                         ])
#                     ]),
#
#             # note: adding color picker for edges i think
#             dcc.Tab(label='Control Panel', children=[
#                 # note: CANT FIND dash reusable components drc so using dcc.Input instead of drc.NamedInput, works
#                 dcc.Dropdown(['prerequisites', 'postrequisites', 'both', 'neither (why would you do this?)'], 'prerequisites', id='reqDropdown1'),
#                 #dcc.Input(name='inColor1', id='inputEdgeColor1', type='text', value='#0074D9'),
#                 #dcc.Input(name='outColor1', id='outputEdgeColor1', type='text', value='#FF4136'),
#             ])
#
#         ])
#     ]),
#     # dhtml.Button('Clear Selection', id='clear-sel-button', n_clicks_timestamp=0),  #FIXME DOESNT WORK - maybe just remove lol
#     # dhtml.Div(id='placeholder')  # note: WTF IS placeholder, no idea what this line is
# ])


def nodeBFSTracer(rootNodeID, direction):
    if direction == "bck":
        edgeEnd = 'target'
        newPtr = 'source'
    elif direction == "fwd":
        edgeEnd = 'source'
        newPtr = 'target'
    else:
        return "invalid direction input"
    finalNodeSet = []
    finalEdgeSet = []
    frontierQ = Queue(maxsize=0)
    frontierQ.put(rootNodeID)
    while not frontierQ.empty():
        curNode = frontierQ.get()
        finalNodeSet.append(curNode)
        curChildren = []
        for eaEdge in myEdges:
            if eaEdge['data'][edgeEnd] == curNode:
                finalEdgeSet.append(eaEdge)
                for eaNode in myNodes:
                    if eaNode['data']['id'] == eaEdge['data'][newPtr] and eaNode['data']['id'] not in finalNodeSet:
                        curChildren.append(eaNode['data']['id'])
        for c in curChildren:
            frontierQ.put(c)
    return finalNodeSet, finalEdgeSet  # note: NodeSet is list of node IDs as str, EdgeSet is JSON OBJECTS!!


# for node click, draw per/post reqs
@myApp.callback(Output(myCyto_id, 'stylesheet'),
                [Input(myCyto_id, 'tapNodeData'),
                 # Input('inputEdgeColor1', 'value'),
                 # Input('outputEdgeColor1', 'value'),
                 Input('reqDropdown1', 'value')])
def generate_stylesheet(node, requisiteDisplayChoice):
    global resetNodeSelection
    # print(resetNodeSelection)
    if not node:
        return myDefaultStylesheet
    if resetNodeSelection:
        resetNodeSelection = False
        return myDefaultStylesheet
    else:
        stylesheet = [{"selector": 'node',
                       'style': {'opacity': 0.3, 'text-justification': 'center', 'text-halign': 'center',
                                 'text-valign': 'center'
                                 }
                       },
                      {'selector': 'edge',
                       'style': {'opacity': 0.2, "curve-style": "bezier",
                                 }
                       },
                      {"selector": 'node[id = "CLER001"]', 'style': {'shape': 'rectangle', 'width': 75, 'height': 40}},

                      # {  # "selector": 'node[id = "{}"]'.format(node['data']['id']),
                      #     "selector": '[title *= a]',
                      #     "style": {  # note: FIXED - (don't) NEED TO KEEP THIS CHUNK EVEN IF IT DOESNT DO ANYTHING BC WEIRD ERROR
                      #         # 'background-color': '#B10DC9',
                      #         # "border-color": "purple",
                      #         # "border-width": 2,
                      #         # "border-opacity": 1,
                      #         # "opacity": 1,
                      #         #
                      #         # "label": "data(label)",
                      #         # "color": "#B10DC9",
                      #         # "text-opacity": 1,
                      #         # "font-size": 12,
                      #         # 'z-index': 9999
                      #     }
                      # }
                      ]
        if requisiteDisplayChoice == 'both':

            prereqNodes, prereqEdges = nodeBFSTracer(node['id'], "fwd")  # FORWARD REQS

            #note: 3lines below JUST FOR TESTING, will separate in dif IF statements for interactive viz options
            bprNodes, bprEdges = nodeBFSTracer(node['id'], "bck")   # BACKWARD REQS
            prereqNodes += bprNodes
            prereqEdges += bprEdges

        elif requisiteDisplayChoice == 'postrequisites':
            prereqNodes, prereqEdges = nodeBFSTracer(node['id'], "fwd")  # FORWARD REQS
        elif requisiteDisplayChoice == 'prerequisites':
            prereqNodes, prereqEdges = nodeBFSTracer(node['id'], "bck")  # BACKWARD REQS
        elif requisiteDisplayChoice == 'neither (why would you do this?)':
            prereqNodes, prereqEdges = [node], []  # empty, for testing

            curNodeID = node['id']  #NOTE this stuff is required for correct formatting without selection
            stylesheet.append({"selector": 'node[id ="{}"]'.format(curNodeID),
                               "style": {'background-color': '#2c4c96',
                                         "border-color": "purple",
                                         "border-width": 2,
                                         "border-opacity": 1,
                                         "opacity": 1,

                                         "color": "#fafbfc",
                                         "text-opacity": 1,
                                         "font-size": 18,
                                         'z-index': 9999}
                               })
        else:
            return "ERROR WITH variable requisiteDisplayChoice - not one of both, prerequisites, postrequisites"

        for eaN in prereqNodes:
            if '%' in eaN:
                stylesheet.append({"selector": 'node[id ="{}"]'.format(eaN),
                                   "style": {'background-color': '#2c4c96',
                                             "border-color": "blue",
                                             "border-width": 5.5,
                                             "border-opacity": 1,
                                             "opacity": .55,  # note: changed-ORs faded

                                             "color": "#fafbfc",
                                             "text-opacity": .75,  # note: changed-ORs faded
                                             "font-size": 18,
                                             'z-index': 9999}
                                   })
            elif '&' in eaN:
                stylesheet.append({"selector": 'node[id ="{}"]'.format(eaN),
                                   "style": {'background-color': '#2c4c96',
                                             "border-color": "red",
                                             "border-width": 6,
                                             "border-opacity": .9,
                                             "opacity": .75,  # note: changed-ANDs faded, less than ORs tho

                                             "color": "#fafbfc",
                                             "text-opacity": .75,  # note: changed-ANDs faded
                                             "font-size": 18,
                                             'z-index': 9999}
                                   })

            else:
                stylesheet.append({"selector": 'node[id ="{}"]'.format(eaN),
                                   "style": {'background-color': '#2c4c96',
                                             "border-color": "purple",
                                             "border-width": 2,
                                             "border-opacity": 1,
                                             "opacity": 1,

                                             "color": "#fafbfc",
                                             "text-opacity": 1,
                                             "font-size": 18,
                                             'z-index': 9999}
                                   })
        for eaE in prereqEdges:
            eaEID = eaE['data']['id']
            if '%' in eaE['data']['target']:
                stylesheet.append({"selector": 'edge[id ="{}"]'.format(eaEID),
                                   "style": {'background-color': '#2c4c96',
                                             "border-color": "blue",
                                             "border-width": 2,
                                             "border-opacity": 1,
                                             "opacity": .75,

                                             "color": "#fafbfc",
                                             "text-opacity": 1,
                                             "font-size": 18,
                                             'z-index': 9999}
                                   })
            else:
                stylesheet.append({"selector": 'edge[id ="{}"]'.format(eaEID),
                                   "style": {'background-color': '#2c4c96',
                                             "border-color": "purple",
                                             "border-width": 2,
                                             "border-opacity": 1,
                                             "opacity": 1,

                                             "color": "#fafbfc",
                                             "text-opacity": 1,
                                             "font-size": 18,
                                             'z-index': 9999}
                                   })

        stylesheet.append(
            {'selector': 'node', 'style': {'label': 'data(id)', 'shape': 'rectangle', 'width': 100, 'height': 60}})
        stylesheet.append({'selector': 'edge', 'style': {'curve-style': 'bezier'}})
        stylesheet.append({'selector': 'edge', 'style': {'mid-target-arrow-color': 'blue',
                                                         'mid-target-arrow-shape': 'vee',
                                                         'line-color': 'grey', 'arrow-scale': 3.5, }}, )
        stylesheet.append(
            {"selector": 'node[id = "CLER001"]', 'style': {'shape': 'rectangle', 'width': 100, 'height': 60}})
        stylesheet.append({'selector': 'node[id *= "%"]',
                           'style': {'label': orLabel}})
        # note: default stylesheet below, copied and appended to new stylesheet cuz i had to readd for some reason

        return stylesheet


# for node click, pulls node data
@myApp.callback(Output('tap-node-data-output1', 'children'),
                [Input(myCyto_id, 'tapNodeData')])
def displayTapNodeData(data):
    if data == None:
        return "select a node to display ORC information here"
    return str(json.dumps(data, indent=2)).strip("{").strip("}")  # indent=2 ?


# for hover over node, pulls node name/title basic info
@myApp.callback(Output('cytoscape-mouseoverNodeData-output', 'children'),
                Input(myCyto_id, 'mouseoverNodeData'))
def displayTapNodeData(data):
    global resetNodeSelection
    if data:
        if data['id'] == 'CLER001':
            resetNodeSelection = True
        return "Hovered Node: " + data['id'] + ": " + data['title']



# does dropdown layout menu stuff
# @myApp.callback(Output('cytoscape-update-layout', 'layout'),
#                 Input('dropdown-update-layout', 'value'))
# def update_layout(layout):
#     return {'name': layout, 'animate': True}


if __name__ == '__main__':
    myApp.run_server(debug=True)
