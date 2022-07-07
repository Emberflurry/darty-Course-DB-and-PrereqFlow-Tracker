import openpyxl
import dash
import dash_cytoscape as dcyto
from dash import html as dhtml

"""Section 1: read from excel to txt w ea line
 a different edge (or node)"""

def xl2SIFnetworkcreator(xlWbFilePath, sheetIndex, startingRowOfEdgeEntries, columnNumofEdgeEntries,
                         outputSIFnameAndOrPath):
    myWB = openpyxl.load_workbook(xlWbFilePath)
    mySheet = myWB[myWB.sheetnames[sheetIndex]]
    with open(outputSIFnameAndOrPath, 'w') as myOutFile:
        glbOrCtr = 0  # can handle up to 100 or nodes per course
        glbAndCtr = 0  # ^

        # FOR outputs to node/edge lists for Dash-Cyto graph visuals, see lines 102-119
        edgeList = []
        cumNodeSet = set()  # for keeping track of nodes to add/already been added
        nodeList = []

        for i in range(startingRowOfEdgeEntries, mySheet.max_row):
            prContents = str(mySheet.cell(row=i, column=columnNumofEdgeEntries).value).strip().split(",")
            if str(prContents) != "['None']":
                orDict = {}  # RESETS FOR EACH COURSE, AVOIDING OVERLAPS
                andDict = {}  # ^
                for ea in prContents:
                    # print("cont " + str(ea))
                    # handle OR node renaming
                    if "%" in ea:
                        curOR_ID = ea[ea.find("%"):ea.find("%") + 3]
                        # print("cur " + str(curOR_ID))
                        if curOR_ID not in orDict:
                            # add new mapping {line% : global%} in orDict
                            if glbOrCtr >= 10:
                                orDict[curOR_ID] = "or" + str(glbOrCtr)
                            elif glbOrCtr < 10:
                                orDict[curOR_ID] = "or0" + str(glbOrCtr)
                            # increment orCounter
                            glbOrCtr += 1
                            # print("ctr" + str(glbOrCtr))

                        if "%" in ea[3:]:  # if OR->OR, also remap second instance
                            # print("2nd OR found")
                            subs = ea[3:]
                            curOR_ID2 = subs[subs.find("%"):subs.find("%") + 3]
                            # print(curOR_ID2)
                            if curOR_ID2 not in orDict:
                                # add new mapping {line% : global%} in orDict
                                if glbOrCtr >= 10:
                                    orDict[curOR_ID2] = "or" + str(glbOrCtr)
                                elif glbOrCtr < 10:
                                    orDict[curOR_ID2] = "or0" + str(glbOrCtr)
                                # increment orCounter
                                glbOrCtr += 1
                                # print("ctr" + str(glbOrCtr))

                        # EITHER WAY set/replace to global mapping (new or existing mapping)
                        ea = ea.replace(curOR_ID, orDict[curOR_ID])
                        ea = ea.replace(curOR_ID2, orDict[curOR_ID2])

                    if "&" in ea:
                        curAND_ID = ea[ea.find("&"):ea.find("&") + 3]
                        # print("cur " + str(curAND_ID))
                        if curAND_ID not in andDict:
                            # add new mapping {line% : global%} in andDict
                            if glbAndCtr >= 10:
                                andDict[curAND_ID] = "and" + str(glbAndCtr)
                            elif glbAndCtr < 10:
                                andDict[curAND_ID] = "and0" + str(glbAndCtr)
                            # increment andCounter
                            glbAndCtr += 1
                            # print("ctr" + str(glbAndCtr))

                        if "&" in ea[3:]:  # if AND->AND, also remap second instance
                            # print("2nd AND found")
                            subs = ea[3:]
                            curAND_ID2 = subs[subs.find("&"):subs.find("&") + 3]
                            # print(curAND_ID2)
                            if curOR_ID2 not in orDict:
                                # add new mapping {line% : global%} in andDict
                                if glbAndCtr >= 10:
                                    andDict[curAND_ID2] = "and" + str(glbAndCtr)
                                elif glbAndCtr < 10:
                                    andDict[curAND_ID2] = "and0" + str(glbAndCtr)
                                # increment andCounter
                                glbAndCtr += 1
                                # print("ctr" + str(glbAndCtr))

                        # EITHER WAY set/replace to global mapping (new or existing mapping)
                        ea = ea.replace(curAND_ID, andDict[curAND_ID])
                        ea = ea.replace(curAND_ID2, andDict[curAND_ID2])

                    # add the lines to the output files for manual checking
                    ea = ea.strip() # clean
                    myOutFile.write(str(ea))
                    myOutFile.write('\n')

                    # split the line data, then add to the list complex  ( ('src','targ','L'), (etc), (etc)...)


                    a = ea.split(" ")
                    # print("A: ")
                    # print(a)
                    mySrc = a[0]
                    if len(a) == 3:
                        myEdgeLabel = a[1]
                        myTarg = a[2]
                        edgeList.append((mySrc,myTarg,myEdgeLabel))

                        if myTarg not in cumNodeSet:
                            c = (myTarg, myTarg)  # TODO CHANGE SECOND ELEM TO SUM USEFUL LATER??
                            nodeList.append(c)
                            cumNodeSet.add(myTarg)

                    if mySrc not in cumNodeSet:
                        b = (mySrc,mySrc)  # TODO CHANGE SECOND ELEM TO SUM USEFUL LATER??
                        nodeList.append(b)
                        cumNodeSet.add(mySrc)

    # print(nodeList)
    # print(edgeList)
    return nodeList, edgeList

cprqfile = "C:/Users/John DeForest/PycharmProjects/dartyclassdb1/2IntermediateProcessing/xlDBcleaning/deleteTestExportCURRENT3.xlsx"
# xl2SIFnetworkcreator(cprqfile, 0, 2, 12, 'testExp.sif')
# print(xl2SIFnetworkcreator(cprqfile, 0, 2, 7, 'ME2.txt'))
myNodesLoL,myEdgesLoL = xl2SIFnetworkcreator(cprqfile, 0, 2, 7, 'ME2.txt')  # 2nd param: 0 for MATH, 1 for MATH+ENGS
# TODO: this txt writing step^ is for manual checking of the reading from excel process,
#  really can just write straight to list format (as is DONE by the fn)

# print("--")
# print(myNodesLoL)
# print(myEdgesLoL)

"""Section2: testing dash-cyto app"""

myApp = dash.Dash(__name__)
myNodes = [
    {'data': {'id': shortID, 'label': labelID1}, }
    for shortID, labelID1 in myNodesLoL
]
myEdges = [
    {'data': {'source': sourceID, 'target': targetID, 'label': labelID2}}
    for sourceID, targetID, labelID2 in myEdgesLoL
]

# print("i--oi")
# print(myNodes)
# print(myEdges)

myAllElements = myNodes + myEdges
myDefaultStylesheet = [
            {'selector': 'node', 'style': {'label': 'data(id)'}},
            {'selector': 'edge', 'style': {'label': 'data(label)'}},
            {'selector': 'edge', 'style': {'curve-style': 'bezier'}},
            {'selector': 'edge', 'style': {'mid-target-arrow-color': 'blue', 'mid-target-arrow-shape': 'vee', 'line-color': 'blue', 'arrow-scale': 3, }}
]

myApp.layout = dhtml.Div([
    dcyto.Cytoscape(
        id='cytoscape',
        elements=myAllElements,
        stylesheet=myDefaultStylesheet,
        style={'width': '100%', 'height': '800px'},
        layout={'name': 'breadthfirst',  # 'cose' for physics layout
                'roots': '[id = "MATH001"]'}
    )
])

if __name__ == '__main__':
    myApp.run_server(debug=True)

