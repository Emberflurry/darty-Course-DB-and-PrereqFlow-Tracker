import openpyxl
import dash
import dash_cytoscape as dcyto
from dash import html as dhtml
# import dash_core_components as dcc # PHASED OUT DO NOT USE THIS IMPORT SYNTAX
from dash import dcc  # ^ use this instead
from dash.dependencies import Input, Output, State
import re
import json
#suppress_callback_exceptions=True
dcyto.load_extra_layouts()
global resetNodeSelection
resetNodeSelection=False

def xl2SIFnetworkcreator(xlWbFilePath, sheetIndex, startingRowOfEdgeEntries, columnNumofEdgeEntries,
                         columnNumofNodeNames, columnNumofTitle, columnNumofDesc,
                         outputSIFnameAndOrPath):
    myWB = openpyxl.load_workbook(xlWbFilePath)
    mySheet = myWB[myWB.sheetnames[sheetIndex]]
    with open(outputSIFnameAndOrPath, 'w') as myOutFile:
        glbOrCtr = 0  # can handle note: INCREASED TO 0-999 (THREE DIGITS CAPACITY FOR ALL GLOBAL ORS/ANDS)
        glbAndCtr = 0  # ^           # note: Only use now if counting # of ors and ands for debugging...

        # FOR outputs to node/edge lists for Dash-Cyto graph visuals, see lines 102-119
        edgeList = []
        nodeList = []

        cumNodeSet = set()  # note: for keeping track of NODES to add/already been added
        lineDuplicateSet = set()  # note: makes sure repeated/duplicate EDGES are not added

        # note: adding nodes from the EDGELISTS
        for i in range(startingRowOfEdgeEntries, mySheet.max_row):
            prContents = str(mySheet.cell(row=i, column=columnNumofEdgeEntries).value).strip().split(",")

            if str(prContents) != "['None']":

                for ea in prContents:
                    if str(ea) not in lineDuplicateSet:
                        lineDuplicateSet.add(ea)

                        if "%" in ea:  # NOTE: the id of the OR nodes will be the %### but ill make the labels all OR s
                            glbOrCtr += 1
                            if "%" in ea[4:]:  # if OR->OR, also remap second instance 3: for 2dig, 4: for 3dig?
                                glbOrCtr += 1

                        if "&" in ea:
                            glbAndCtr += 1
                            if "&" in ea[4:]:  # if AND->AND, also remap second instance
                                glbAndCtr += 1

                        # add the lines to the output files for manual checking
                        ea = ea.strip()  # clean
                        myOutFile.write(str(ea))
                        myOutFile.write('\n')

                        # split the line data, then add to the list complex  ( ('src','targ','L'), (etc), (etc)...)

                        a = ea.split(" ")
                        mySrc = a[0]
                        if len(a) == 3:
                            myEdgeLabel = a[1]
                            myTarg = a[2]
                            edgeList.append((mySrc, myTarg, myEdgeLabel))

                            if myTarg not in cumNodeSet:
                                c = (myTarg, "", "")
                                nodeList.append(c)
                                cumNodeSet.add(myTarg)

                        if mySrc not in cumNodeSet:
                            b = (mySrc, "", "")
                            nodeList.append(b)
                            cumNodeSet.add(mySrc)

        # note: now add FLOATER NODES - ones that have no prereqs/connections
        for k in range(startingRowOfEdgeEntries, mySheet.max_row):
            nodeInfo = str(mySheet.cell(row=k, column=columnNumofNodeNames).value).strip()
            courseTitle = str(mySheet.cell(row=k, column=columnNumofTitle).value).strip()
            courseDescription = str(mySheet.cell(row=k, column=columnNumofDesc).value).strip()
            # TODO: ALSO ADD the OG PREREQ INFO (TEXT) SO THAT USERS CAN INTERPRET COMPLEXITIES
            # AND OFC the URL(SEE recent ORC scrape xlsx in Webscrapers folder and corresponding script-
            # ^integrate into final version),
            # coreqs,profs, etc -all that can be found from orc page...
            # maybe later also link relevant DEPARTMENT pages but that might be manual

            if "." not in nodeInfo:
                match = re.match(r"([a-z]+)([0-9]+)", nodeInfo, re.I)
                if match:
                    splitNodeInfo = match.groups()
                    # print(splitNodeInfo)

                    if len(splitNodeInfo[1]) == 1:
                        newNumStr = "00" + splitNodeInfo[1]
                    elif len(splitNodeInfo[1]) == 2:
                        newNumStr = "0" + splitNodeInfo[1]
                    else:
                        newNumStr = splitNodeInfo[1]

                    floaterNode = splitNodeInfo[0] + newNumStr
                else:
                    print("regex error in finding match of char/number boundary")

            elif "." in nodeInfo:
                floaterNode = nodeInfo

            newNodeTuple = floaterNode, courseTitle, courseDescription
            print("new", newNodeTuple)

            duplicateReplaced = False
            for g in range(0, len(nodeList) - 1):
                if floaterNode == nodeList[g][0]:
                    nodeList[g] = newNodeTuple
                    duplicateReplaced = True
                    break
            if duplicateReplaced is False:
                nodeList.append(newNodeTuple)

    # print(nodeList)
    # print(edgeList)
        for y in range(0, len(nodeList)-1):
            if nodeList[y][0][0] == "%":
                newOrTitle = "OR - 1 of child nodes is required to be fulfilled as prerequisite"
                nodeList[y] = nodeList[y][0], newOrTitle, "just an OR node"
            if nodeList[y][0][0] == "&":
                newAndTitle = "AND - all children are required to be fulfilled as prerequisite"
                nodeList[y] = nodeList[y][0], newAndTitle, "just an AND node"

    return nodeList, edgeList


cprqfile = "C:/Users/John DeForest/PycharmProjects/dartyclassdb1/2IntermediateProcessing/xlDBcleaning/deleteTestExportCURRENT3.xlsx"

myNodesLoL, myEdgesLoL = xl2SIFnetworkcreator(cprqfile, 0, 2, 7, 1, 2, 5,
                                              'edgelistOutput3.txt')  # 2nd param: 0 for MATH, 1 for MATH+ENGS
# note: this txt writing step^ is for manual checking of the reading from excel process,
#  really can just write straight to list format (as is DONE by the fn)

# print("--")
# print(myNodesLoL)
# print(myEdgesLoL)
# quit()

myApp = dash.Dash(__name__)

myNodes = [
    {'data': {'id': shortID, 'title': labelID1, 'desc': descText}, }
    for shortID, labelID1, descText in myNodesLoL
]
myEdges = [
    {'data': {'source': sourceID, 'target': targetID, 'label': labelID2}}
    for sourceID, targetID, labelID2 in myEdgesLoL
]

# for eaNode in myNodes:
#     if eaNode['data']['id']=="MATH022":
#         m22 = eaNode['data']['id']
#print(m22)
from queue import Queue
nodeSet = []
myQ = Queue(maxsize=0)
myQ.put('MATH086')
while not myQ.empty():
    node = myQ.get()
    nodeSet.append(node)

    children = []
    for e in myEdges:
        if e['data']['target'] == node:
            for k in myNodes:
                if k['data']['id'] == e['data']['source'] and k['data']['id'] not in nodeSet:
                    children.append(k['data']['id'])
    for c in children:
        myQ.put(c)

print(nodeSet)

def nodeBacktracer(rootNodeID):
    finalNodeSet=[]
    finalEdgeSet=[]
    frontierQ=Queue(maxsize=0)
    frontierQ.put(rootNodeID)
    while not frontierQ.empty():
        curNode=frontierQ.get()
        finalNodeSet.append(curNode)
        curChildren=[]
        for eaEdge in myEdges:
            if eaEdge['data']['target'] == curNode:
                finalEdgeSet.append(eaEdge)
                for eaNode in myNodes:
                    if eaNode['data']['id'] == eaEdge['data']['source'] and eaNode['data']['id'] not in finalNodeSet:
                        curChildren.append(eaNode['data']['id'])
        for c in curChildren:
            frontierQ.put(c)
    return finalNodeSet,finalEdgeSet