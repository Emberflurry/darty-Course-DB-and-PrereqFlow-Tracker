import openpyxl
import dash
import dash_cytoscape as dcyto
from dash import html as dhtml
from dash import dcc  # ^ use this instead
from dash.dependencies import Input, Output, State
import re
import json
from queue import Queue


dcyto.load_extra_layouts()


# file is excel file to load from
# sheet is name of the sheet in the file
def get_graph_data(file, sheet):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]

    # 
    edges = set()
    nodes = dict()

    # Starts at 2 to skip title row and b/c openpyxl is 1-based instead of 0-based
    for i in range(2, sheet.max_row+1):
        
        # Get node info

        # Column 1 is course name
        nodeName = str(sheet.cell(row=i, column=1).value).strip()
        # Column 2 is course title
        nodeTitle = str(sheet.cell(row=i, column=2).value).strip()
        # Column 5 is the description
        nodeDescription = str(sheet.cell(row=i, column=5).value).strip()

        # Rewrite name with leading 0s lol
        if nodeName != "None":
            if not "." in nodeName:
                # segments into (department, number)
                segmented = re.match(r'([a-z]+)([0-9]+)', nodeName, re.I)
                # makes sure there was a match
                if segmented:
                    # Rewrite the name with leading zeros for 3 digits
                    dep, number = segmented.groups()
                    nodeName = dep + number.zfill(3)

            nodes[nodeName] = {"title":nodeTitle, "disc":nodeDescription}


        # Get edge info

        edges_data = str(sheet.cell(row=i, column=7).value).strip().split(",")

        if edges_data != ["None"]:
            for edge_data in edges_data:
                new_edge = edge_data.split(" ")
                assert len(new_edge) == 3

                startNode = new_edge[0]
                label = new_edge[1]
                endNode = new_edge[2]
                
                if not startNode in nodes.keys():
                    nodes[startNode] = {"title":None, "disc":None}

                if not endNode in nodes.keys():
                    nodes[endNode] = {"title":None, "disc":None}

                # add edge
                edges.add((startNode, endNode, label))


    return nodes, edges

file = "../../2IntermediateProcessing/xlDBcleaning/deleteTestExportCURRENT3.xlsx"

nodes, edges = get_graph_data(file, "MATHglobalsRedo")

app = dash.Dash(__name__, title='Darty Course-Flow Vizualizer')

defaultStylesheet = [
    {'selector': 'node',
     'style': {'label': 'data(id)', 'shape': 'rectangle', 'width': 100, 'height': 50, 'text-justification': 'center',
               'text-halign': 'center',
               'text-valign': 'center'}},
    {'selector': 'edge', 'style': {'curve-style': 'bezier'}},
    {'selector': 'edge', 'style': {'mid-target-arrow-color': 'blue',
                                   'mid-target-arrow-shape': 'vee',
                                   'line-color': 'grey', 'arrow-scale': 3.5, }},
    {"selector": 'node[id = "CLER001"]', 'style': {'shape': 'rectangle', 'width': 100, 'height': 60}},
    {'selector': 'node[id *= "%"]',
                             'style': {'label': "OR"}}
]

sidebarStyles = {
    'tab1': {'height': 'calc(98vh-115px)'},
    'contentStyle1': {'overflow-y': 'scroll', 'height': 'calc(50%-25px)', 'border': 'thin lightgrey solid'}
}

app.layout = dhtml.Div([
    dhtml.H1(children='Dartmouth Course-Flow Vizualizer', style={'color': '#096A3F'}),
    dhtml.Div(children="course data from ORC as of 7/22, developed and maintained by John DeForest and visualized through a Plotly-Dash-Cytoscape config"),
    # note: this div for cyto node layout
    dhtml.Div(className='eight columns', children=[
        dcyto.Cytoscape(
            id=myCyto_id,
            elements=myAllElements,
            stylesheet=myDefaultStylesheet,
            style={'width': '90%', 'height': '450px'},
            layout={'name': 'dagre',
                    'roots': '[id = "MATH001"]'}
        ),
        dhtml.Blockquote(id='cytoscape-mouseoverNodeData-output')
    ]),

    # note: this div 4 sidebar info
    # TODO: MAKE INTO AN ACTUAL SIDEBAR, NOT THE BOTTOM TAB DISPLAY
    dhtml.Div(className='four columns', children=[
        # dcc.Textarea #TODO try this instead/other html content types next after testing 'Tabs' as below
        dcc.Tabs(id='tabs', children=[
            dcc.Tab(label='Click on a node for full ORC content',  # label here is the TITLE of the info tab
                    children=[
                        dhtml.Div(style=sidebarStyles['tab1'], children=[
                            dhtml.P('Node Data:'),  # SUBTITLE just above content
                            dhtml.Pre(
                                id='tap-node-data-output1',
                                style=sidebarStyles['contentStyle1']
                            )
                        ])
                    ]),

            # note: adding color picker for edges i think
            dcc.Tab(label='Control Panel', children=[
                # note: CANT FIND dash reusable components drc so using dcc.Input instead of drc.NamedInput, works
                dcc.Dropdown(['prerequisites', 'postrequisites', 'both', 'neither (why would you do this?)'], 'prerequisites', id='reqDropdown1'),
                #dcc.Input(name='inColor1', id='inputEdgeColor1', type='text', value='#0074D9'),
                #dcc.Input(name='outColor1', id='outputEdgeColor1', type='text', value='#FF4136'),
            ])

        ])
    ]),
    # dhtml.Button('Clear Selection', id='clear-sel-button', n_clicks_timestamp=0),  #FIXME DOESNT WORK - maybe just remove lol
    # dhtml.Div(id='placeholder')  # note: WTF IS placeholder, no idea what this line is
])



"""

# for node click, draw per/post reqs
@myApp.callback(Output(myCyto_id, 'stylesheet'),
                [Input(myCyto_id, 'tapNodeData'),
                 # Input('inputEdgeColor1', 'value'),
                 # Input('outputEdgeColor1', 'value'),
                 Input('reqDropdown1', 'value')])
def generate_stylesheet(node, requisiteDisplayChoice):
    global resetNodeSelection
    # print(resetNodeSelection)
    if not node:
        return myDefaultStylesheet
    if resetNodeSelection:
        resetNodeSelection = False
        return myDefaultStylesheet
    else:
        stylesheet = [{"selector": 'node',
                       'style': {'opacity': 0.3, 'text-justification': 'center', 'text-halign': 'center',
                                 'text-valign': 'center'
                                 }
                       },
                      {'selector': 'edge',
                       'style': {'opacity': 0.2, "curve-style": "bezier",
                                 }
                       },
                      {"selector": 'node[id = "CLER001"]', 'style': {'shape': 'rectangle', 'width': 75, 'height': 40}},

                      # {  # "selector": 'node[id = "{}"]'.format(node['data']['id']),
                      #     "selector": '[title *= a]',
                      #     "style": {  # note: FIXED - (don't) NEED TO KEEP THIS CHUNK EVEN IF IT DOESNT DO ANYTHING BC WEIRD ERROR
                      #         # 'background-color': '#B10DC9',
                      #         # "border-color": "purple",
                      #         # "border-width": 2,
                      #         # "border-opacity": 1,
                      #         # "opacity": 1,
                      #         #
                      #         # "label": "data(label)",
                      #         # "color": "#B10DC9",
                      #         # "text-opacity": 1,
                      #         # "font-size": 12,
                      #         # 'z-index': 9999
                      #     }
                      # }
                      ]
        if requisiteDisplayChoice == 'both':

            prereqNodes, prereqEdges = nodeBFSTracer(node['id'], "fwd")  # FORWARD REQS

            #note: 3lines below JUST FOR TESTING, will separate in dif IF statements for interactive viz options
            bprNodes, bprEdges = nodeBFSTracer(node['id'], "bck")   # BACKWARD REQS
            prereqNodes += bprNodes
            prereqEdges += bprEdges

        elif requisiteDisplayChoice == 'postrequisites':
            prereqNodes, prereqEdges = nodeBFSTracer(node['id'], "fwd")  # FORWARD REQS
        elif requisiteDisplayChoice == 'prerequisites':
            prereqNodes, prereqEdges = nodeBFSTracer(node['id'], "bck")  # BACKWARD REQS
        elif requisiteDisplayChoice == 'neither (why would you do this?)':
            prereqNodes, prereqEdges = [node], []  # empty, for testing

            curNodeID = node['id']  #NOTE this stuff is required for correct formatting without selection
            stylesheet.append({"selector": 'node[id ="{}"]'.format(curNodeID),
                               "style": {'background-color': '#2c4c96',
                                         "border-color": "purple",
                                         "border-width": 2,
                                         "border-opacity": 1,
                                         "opacity": 1,

                                         "color": "#fafbfc",
                                         "text-opacity": 1,
                                         "font-size": 18,
                                         'z-index': 9999}
                               })
        else:
            return "ERROR WITH variable requisiteDisplayChoice - not one of both, prerequisites, postrequisites"

        for eaN in prereqNodes:
            if '%' in eaN:
                stylesheet.append({"selector": 'node[id ="{}"]'.format(eaN),
                                   "style": {'background-color': '#2c4c96',
                                             "border-color": "blue",
                                             "border-width": 5.5,
                                             "border-opacity": 1,
                                             "opacity": .55,  # note: changed-ORs faded

                                             "color": "#fafbfc",
                                             "text-opacity": .75,  # note: changed-ORs faded
                                             "font-size": 18,
                                             'z-index': 9999}
                                   })
            elif '&' in eaN:
                stylesheet.append({"selector": 'node[id ="{}"]'.format(eaN),
                                   "style": {'background-color': '#2c4c96',
                                             "border-color": "red",
                                             "border-width": 6,
                                             "border-opacity": .9,
                                             "opacity": .75,  # note: changed-ANDs faded, less than ORs tho

                                             "color": "#fafbfc",
                                             "text-opacity": .75,  # note: changed-ANDs faded
                                             "font-size": 18,
                                             'z-index': 9999}
                                   })

            else:
                stylesheet.append({"selector": 'node[id ="{}"]'.format(eaN),
                                   "style": {'background-color': '#2c4c96',
                                             "border-color": "purple",
                                             "border-width": 2,
                                             "border-opacity": 1,
                                             "opacity": 1,

                                             "color": "#fafbfc",
                                             "text-opacity": 1,
                                             "font-size": 18,
                                             'z-index': 9999}
                                   })
        for eaE in prereqEdges:
            eaEID = eaE['data']['id']
            if '%' in eaE['data']['target']:
                stylesheet.append({"selector": 'edge[id ="{}"]'.format(eaEID),
                                   "style": {'background-color': '#2c4c96',
                                             "border-color": "blue",
                                             "border-width": 2,
                                             "border-opacity": 1,
                                             "opacity": .75,

                                             "color": "#fafbfc",
                                             "text-opacity": 1,
                                             "font-size": 18,
                                             'z-index': 9999}
                                   })
            else:
                stylesheet.append({"selector": 'edge[id ="{}"]'.format(eaEID),
                                   "style": {'background-color': '#2c4c96',
                                             "border-color": "purple",
                                             "border-width": 2,
                                             "border-opacity": 1,
                                             "opacity": 1,

                                             "color": "#fafbfc",
                                             "text-opacity": 1,
                                             "font-size": 18,
                                             'z-index': 9999}
                                   })

        stylesheet.append(
            {'selector': 'node', 'style': {'label': 'data(id)', 'shape': 'rectangle', 'width': 100, 'height': 60}})
        stylesheet.append({'selector': 'edge', 'style': {'curve-style': 'bezier'}})
        stylesheet.append({'selector': 'edge', 'style': {'mid-target-arrow-color': 'blue',
                                                         'mid-target-arrow-shape': 'vee',
                                                         'line-color': 'grey', 'arrow-scale': 3.5, }}, )
        stylesheet.append(
            {"selector": 'node[id = "CLER001"]', 'style': {'shape': 'rectangle', 'width': 100, 'height': 60}})
        stylesheet.append({'selector': 'node[id *= "%"]',
                           'style': {'label': orLabel}})
        # note: default stylesheet below, copied and appended to new stylesheet cuz i had to readd for some reason

        return stylesheet


# for node click, pulls node data
@myApp.callback(Output('tap-node-data-output1', 'children'),
                [Input(myCyto_id, 'tapNodeData')])
def displayTapNodeData(data):
    if data == None:
        return "select a node to display ORC information here"
    return str(json.dumps(data, indent=2)).strip("{").strip("}")  # indent=2 ?


# for hover over node, pulls node name/title basic info
@myApp.callback(Output('cytoscape-mouseoverNodeData-output', 'children'),
                Input(myCyto_id, 'mouseoverNodeData'))
def displayTapNodeData(data):
    global resetNodeSelection
    if data:
        if data['id'] == 'CLER001':
            resetNodeSelection = True
        return "Hovered Node: " + data['id'] + ": " + data['title']



# does dropdown layout menu stuff
# @myApp.callback(Output('cytoscape-update-layout', 'layout'),
#                 Input('dropdown-update-layout', 'value'))
# def update_layout(layout):
#     return {'name': layout, 'animate': True}


if __name__ == '__main__':
    myApp.run_server(debug=True)
"""