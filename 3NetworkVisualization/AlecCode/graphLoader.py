import openpyxl
import re

# file is excel file to load from
# sheet is name of the sheet in the file
def get_graph_data(file, sheet):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]


    edges = list()
    nodes = dict()

    # Starts at 2 to skip title row and b/c openpyxl is 1-based instead of 0-based
    for i in range(2, sheet.max_row+1):
        
        # Get node info

        # Column 1 is course name
        nodeName = str(sheet.cell(row=i, column=1).value).strip()
        # Column 2 is course title
        nodeTitle = str(sheet.cell(row=i, column=2).value).strip()
        # Column 5 is the description
        nodeDescription = str(sheet.cell(row=i, column=5).value).strip()

        # Rewrite name with leading 0s lol
        if nodeName != "None":
            if not "." in nodeName:
                # segments into (department, number)
                segmented = re.match(r'([a-z]+)([0-9]+)', nodeName, re.I)
                # makes sure there was a match
                if segmented:
                    # Rewrite the name with leading zeros for 3 digits
                    dep, number = segmented.groups()
                    nodeName = dep + number.zfill(3)

            nodes[nodeName] = {"name":nodeName, "title":nodeTitle, "desc":nodeDescription}


        # Get edge info

        edges_data = str(sheet.cell(row=i, column=7).value).strip().split(",")

        if edges_data != ["None"]:
            for edge_data in edges_data:
                new_edge = edge_data.split(" ")
                assert len(new_edge) == 3

                startNode = new_edge[0]
                label = new_edge[1]
                endNode = new_edge[2]
                
                if not startNode in nodes.keys():
                    nodes[startNode] = {"name":startNode, "title":None, "desc":None}

                if not endNode in nodes.keys():
                    nodes[endNode] = {"name":endNode, "title":None, "desc":None}

                # add edge
                edge = {"startNode":startNode, "endNode":endNode, "label":label}
                if not edge in edges:
                    edges.append(edge)


    return nodes, edges