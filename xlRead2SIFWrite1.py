import networkx as nx
import plotly.graph_objects as go
from courseclass import Course
import matplotlib.pyplot as plt
import openpyxl

cprqfile = "C:/Users/John DeForest/PycharmProjects/dartyclassdb1/deltestexp.xlsx"
sheet1 = "MathOnly4networkTesting"
# TODO: finish math dept prereq entry in SIF-readable CSV format (IN DELTESTEXP.xlsx), then to all depts(big)
wb_obj = openpyxl.load_workbook(cprqfile)
mathSheet = wb_obj[wb_obj.sheetnames[0]]

with open('testExp.sif', 'w') as myOutFile:
    glbOrCtr = 0  # can handle up to 100 or nodes per course
    glbAndCtr = 0  # ^
    # for i in range(2, mathSheet.max_row):
    while True:
        i = 49
        prContents = str(mathSheet.cell(row=i, column=12).value).strip().split(",")
        if str(prContents) != "['None']":
            orDict = {}  # RESETS FOR EACH COURSE, AVOIDING OVERLAPS
            andDict = {}  # ^
            for ea in prContents:
                print(ea)
                # handle OR node renaming
                if "%" in ea:
                    curOR_ID = ea[ea.find("%"):ea.find("%")+3]
                    print(curOR_ID)
                    if curOR_ID not in orDict:
                        # add new mapping {line% : global%} in orDict
                        if glbOrCtr >= 10:
                            orDict[curOR_ID] = "or"+str(glbOrCtr)
                        elif glbOrCtr < 10:
                            orDict[curOR_ID] = "or0"+str(glbOrCtr)
                        # increment orCounter
                        glbOrCtr += 1
                        print(glbOrCtr)
                    # EITHER WAY set/replace to global mapping (new or existing mapping)
                    ea = ea.replace(curOR_ID, orDict[curOR_ID])

                    if ea[ea.find("%")+3:].find("%") != -1:  #TODO: if OR1 -> OR2, need to run replacer again:
                        curOR_ID = ea[ea[ea.find("%")+3:].find("%") : ea[ea.find("%")+3:].find("%")+3]
                        print(curOR_ID)
                        if curOR_ID not in orDict:
                            # add new mapping {line% : global%} in orDict
                            if glbOrCtr >= 10:
                                orDict[curOR_ID] = "or" + str(glbOrCtr)
                            elif glbOrCtr < 10:
                                orDict[curOR_ID] = "or0" + str(glbOrCtr)
                            # increment orCounter
                            glbOrCtr += 1
                            print(glbOrCtr)
                        # EITHER WAY set/replace to global mapping (new or existing mapping)
                        ea = ea.replace(curOR_ID, orDict[curOR_ID])

                # handle AND node renaming
                if "&" in ea:
                    curAND_ID = ea[ea.find("&"):ea.find("&")+3]
                    print(curAND_ID)

                    if curAND_ID not in andDict:
                        # add new mapping {line% : global%} in andDict
                        if glbAndCtr >= 10:
                            andDict[curAND_ID] = "and"+str(glbAndCtr)
                        elif glbAndCtr < 10:
                            andDict[curAND_ID] = "and0"+str(glbAndCtr)
                        # increment andCounter
                        glbAndCtr += 1
                        print(glbAndCtr)

                    # EITHER WAY set/replace to global mapping (new or existing mapping)
                    ea = ea.replace(curAND_ID, andDict[curAND_ID])

                myOutFile.write(str(ea))
                myOutFile.write('\n')
        break

# TODO: change network display settings/etc
# TODO: auto-open in browser/etc, also make interactive (long term, using Dash-Cytoscape.js?)
