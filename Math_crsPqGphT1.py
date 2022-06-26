import networkx as nx
import plotly.graph_objects as go
from courseclass import Course
import matplotlib.pyplot as plt
import openpyxl
cprqfile = "C:/Users/John DeForest/PycharmProjects/dartyclassdb1/deltestexp.xlsx"
sheet1 = "MathOnly4networkTesting"
wb_obj = openpyxl.load_workbook(cprqfile)
print(wb_obj.sheetnames)
mathSheet = wb_obj[wb_obj.sheetnames[0]]

hash2coursenameDict = {}
cname2prereqsDict = {}

for i in range(2, 60):
    readHash = mathSheet.cell(row=i, column=9).value
    print(readHash)
    readCourseName = mathSheet.cell(row=i, column=8).value
    print(readCourseName)
    hash2coursenameDict[readHash] = readCourseName

    readCrsPrereqs = mathSheet.cell(row=i, column=10).value
    print(readCrsPrereqs)
    if readCrsPrereqs == None:
        cname2prereqsDict[readCourseName] = ""
    else:
        cname2prereqsDict[readCourseName] = readCrsPrereqs

print(hash2coursenameDict)
print(cname2prereqsDict)
print(hash2coursenameDict["0015"])
print(cname2prereqsDict["MATH013"])
# quit()


def get_key_by_value(find_val, dict):
    for key, value in dict.items():
        if find_val == value:
            return key
    return "key4: " + find_val + " DNE"


myG = nx.DiGraph()
for i in hash2coursenameDict:
    myG.add_node(hash2coursenameDict[i])
    # G.add_node(i)
# nx.draw(G, with_labels=True, font_weight='bold')
# plt.show()



global orCounter, andCounter  # bc the OR nodes need to have unique IDs
orCounter = 0
andCounter = 0


def crsgraphcreator(courseHash):
    #TODO 6/25/22 currently cannot handle 2of x level courses, etc for philosophy/other socials classes,
    # maybe test w ors? maybe can do with just ORs and ANDs but test prereqs w type of OR node

    global orCounter, andCounter
    courseName = hash2coursenameDict[courseHash]
    prereqString = cname2prereqsDict[courseName].strip()
    if prereqString != "" and prereqString is not None:
        allPrConditions = prereqString.split("/")
        print(allPrConditions)

        for ea in allPrConditions:
            if ":" in ea:  # OR List
                newOrNode = str("or" + str(orCounter))
                myG.add_node(newOrNode)
                myG.add_edge(newOrNode, courseName)
                orCounter += 1
                orItems = ea.strip().split(":")
                print(orItems)

                for eaPrq in orItems:
                    if "(" and ")" and "+" in eaPrq:
                        newAndNode = str("and"+str(andCounter))
                        myG.add_node(newAndNode)
                        myG.add_edge(newAndNode, newOrNode)
                        andCounter += 1

                        advANDs = eaPrq.strip("(").strip(")").split("+")
                        for q in advANDs:
                            print("q:"+str(q))
                            myG.add_edge(q, newAndNode)

                    else:
                        myG.add_edge(eaPrq, newOrNode)
            elif ":" not in ea and "(" not in ea and ")" not in ea and "+" not in ea:  # single course
                myG.add_edge(ea, courseName)
    else:
        print(courseName + " hasNoPrqs")


for eaCrs in cname2prereqsDict:  # loops over prereq dict
    crsgraphcreator(get_key_by_value(eaCrs, hash2coursenameDict))
# crsgraphcreator("0008")
# crsgraphcreator("0009")
# crsgraphcreator("0002")

nx.draw(myG, with_labels=True, font_weight='bold', node_color='grey')
plt.show()
