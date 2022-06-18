from xlwt import Workbook
import pandas as pd
import xlrd
import openpyxl

#read_file = pd.read_excel (r'C:\Users\John DeForest\PycharmProjects\dartyclassdb1\orcALL_UGScrape3.xls')
#read_file.to_csv (r'C:\Users\John DeForest\PycharmProjects\dartyclassdb1\orcAll_UGScrape3.csv', index = None, header=True)

#orcdataunop = "orcAll_UGScrape3.csv"
#orcdataop = open(orcdataunop)

# wb1 = Workbook()
# sheet1 = wb1.add_sheet('cleaned')
#
#
#
#
#     wb1.save('orcdatacleaned.xls')
#
#
# wb1.save('orcdatacleaned.xls')

# wb1 = xlrd.open_workbook_xls('deltest.xls')
# ws1 = wb1.sheet_by_index(0)
# totrows = ws1.nrows - 1

# i = 0
# while i < totrows:
#     cellval = ws1.cell(i, 0)
#     if cellval == "" or cellval == None:
#         ws1.delete_rows(i, i+1)
#     else:
#         i += 1
openpath = 'deltest.xlsx'
wb1op = openpyxl.load_workbook(openpath)
ws1op = wb1op['rawexport']

# wb1r = xlrd.open_workbook(openpath)
# ws1r = wb1r.sheet_by_index(0)

i = 1
while i < ws1op.max_row:
    if ws1op.cell(i, 1) == "":
        ws1op.delete_rows(i)
    else:
        i += 1
# def remove(sheet):
#     for row in sheet.iter_rows():
#         #if row[0] == "" or row[0] == None:
#         sheet.delete_rows(row[0].row, 1)
#         remove(sheet)
#         # if sheet.cell(row, 0) == "" or sheet.cell(row, 0) == None:
#         #     sheet.delete_rows(row[0].row, 1)
#         #     remove(sheet)
#
# for row in ws1:
#     remove(ws1)

savepath = 'deltestexp.xlsx'
wb1op.save(savepath)


